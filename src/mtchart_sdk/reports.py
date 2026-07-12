from __future__ import annotations

import json
import re
import zipfile
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable
from xml.sax.saxutils import escape

from mtchart_sdk.models import PartItem
from mtchart_sdk.output_paths import build_output_paths, sanitize_path_part
from mtchart_sdk.rules import clean_identifier, normalize_item, total_quantity


@dataclass(frozen=True)
class ReportIdentity:
    report_number: str
    pn: str = ""
    sn: str = ""
    oven: str = ""
    project: str = ""

    @property
    def safe_report_number(self) -> str:
        return format_report_number(self.report_number) or "SEM_RELATORIO"


@dataclass(frozen=True)
class TemperatureLogPoint:
    timestamp: datetime
    value: float
    pen: str = ""


def format_report_number(value: object) -> str:
    return str(value or "").strip().replace("/", "-")


def is_batch_identifier(value: object, batch_label: str = "LOTE") -> bool:
    text = str(value or "").strip().upper()
    label = str(batch_label or "LOTE").upper()
    return text.startswith(("LOTE", "BATCH", label))


def serial_display_for_mass(sn: object, items_mass: str | Iterable[object] | None = None, multiple_label: str = "MULTIPLO") -> str:
    text = str(sn or "").strip()
    if text != multiple_label:
        return text
    if not items_mass:
        return ""
    try:
        items = json.loads(items_mass) if isinstance(items_mass, str) else items_mass
    except (TypeError, ValueError):
        return text
    for item in items or []:
        if isinstance(item, dict) and str(item.get("sn", "")).strip():
            return text
        if not isinstance(item, dict) and str(item).strip():
            return text
    return ""


def format_datetime_display(value: datetime | str | None, lang: str = "PT") -> str:
    if not value:
        return "-"
    if isinstance(value, datetime):
        date_value = value
    else:
        text = str(value).split(".")[0]
        try:
            date_value = datetime.strptime(text, "%Y-%m-%d %H:%M:%S")
        except ValueError:
            return text
    date_format = "%Y/%m/%d %H:%M:%S" if str(lang).upper() == "EN" else "%d/%m/%Y %H:%M:%S"
    return date_value.strftime(date_format)


def format_audit_action(value: object, labels: dict[str, str] | None = None) -> str:
    labels = labels or {}
    action = str(value or "").strip()
    if not action:
        return ""
    return labels.get(f"audit_action_{action.lower()}", action)


def format_audit_tab(value: object, labels: dict[str, str] | None = None) -> str:
    labels = labels or {}
    tab = str(value or "").strip()
    if not tab:
        return ""
    normalized = re.sub(r"[^a-z0-9]+", "_", tab.lower()).strip("_")
    return labels.get(f"audit_tab_{normalized}", tab)


def _format_audit_detail_value(value: object, labels: dict[str, str]) -> str:
    raw = str(value or "").strip()
    lower = raw.lower()
    if lower == "true":
        return labels.get("txt_sim", "SIM")
    if lower == "false":
        return labels.get("txt_nao", "NAO")
    return raw


def format_audit_details(value: object, labels: dict[str, str] | None = None) -> str:
    labels = labels or {}
    details = str(value or "").strip()
    if not details:
        return ""
    phrase_key = f"audit_detail_phrase_{details.lower().replace(' ', '_').replace('.', '')}"
    if phrase_key in labels:
        return labels[phrase_key]
    if "->" in details and ";" not in details and "=" not in details:
        left, right = details.split("->", 1)
        return f"{format_audit_tab(left, labels)} {labels.get('audit_detail_to', '->')} {format_audit_tab(right, labels)}"

    parts = []
    parsed_any = False
    for raw_part in details.split(";"):
        part = raw_part.strip()
        if not part:
            continue
        if "=" not in part:
            parts.append(part)
            continue
        parsed_any = True
        key, raw_val = part.split("=", 1)
        key = key.strip()
        label = labels.get(f"audit_detail_{key.lower()}", key.replace("_", " ").title())
        parts.append(f"{label}: {_format_audit_detail_value(raw_val, labels)}")
    return "; ".join(parts) if parsed_any else details


def format_audit_log_for_display(log: dict[str, object], labels: dict[str, str] | None = None) -> dict[str, object]:
    display = dict(log or {})
    display["aba"] = format_audit_tab(display.get("aba", ""), labels)
    display["acao"] = format_audit_action(display.get("acao", ""), labels)
    display["detalhes"] = format_audit_details(display.get("detalhes", ""), labels)
    return display


def format_audit_logs_for_display(
    logs: Iterable[dict[str, object]],
    labels: dict[str, str] | None = None,
) -> list[dict[str, object]]:
    return [format_audit_log_for_display(log, labels) for log in (logs or [])]


def export_audit_operations(
    logs: Iterable[dict[str, object]],
    output_path: str | Path,
    labels: dict[str, str] | None = None,
    filters: dict[str, object] | None = None,
    generated_at: datetime | None = None,
) -> Path:
    labels = labels or {}
    filters = filters or {}
    rows = list(logs or [])
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.page import PageMargins
    except ImportError:
        return _export_audit_operations_basic_xlsx(rows, output, labels, filters, generated_at)

    wb = Workbook()
    ws = wb.active
    ws.title = labels.get("audit_logs_title", "Operation logs")[:31]
    header_fill = PatternFill(start_color="1F538D", end_color="1F538D", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.append([labels.get("audit_logs_title", "Operation logs")])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    ws["A1"].fill = header_fill
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=14)
    ws["A1"].alignment = center
    ws.append([labels.get("audit_export_generated_at", "Generated at"), format_datetime_display(generated_at or datetime.now())])
    if filters.get("data_inicio") or filters.get("start"):
        ws.append([labels.get("audit_filter_start", "Start date/time"), filters.get("data_inicio") or filters.get("start")])
    if filters.get("data_fim") or filters.get("end"):
        ws.append([labels.get("audit_filter_end", "End date/time"), filters.get("data_fim") or filters.get("end")])
    ws.append([labels.get("audit_export_total", "Total records"), len(rows)])
    ws.append([])

    ws.append([
        labels.get("audit_col_datetime", "Date/time"),
        labels.get("audit_col_operator", "Operator ID"),
        labels.get("audit_col_tab", "Tab"),
        labels.get("audit_col_action", "Action"),
        labels.get("audit_col_details", "Details"),
    ])
    header_row = ws.max_row
    for cell in ws[header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    for log in format_audit_logs_for_display(rows, labels):
        ws.append([
            log.get("data_hora", ""),
            log.get("matricula", ""),
            log.get("aba", ""),
            log.get("acao", ""),
            log.get("detalhes", ""),
        ])
        for cell in ws[ws.max_row]:
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for idx, width in enumerate([22, 16, 22, 28, 60], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    for row_idx in range(2, header_row):
        ws.cell(row=row_idx, column=1).font = Font(bold=True)
    ws.freeze_panes = f"A{header_row + 1}"
    ws.sheet_view.showGridLines = False
    ws.auto_filter.ref = f"A{header_row}:E{max(header_row, ws.max_row)}"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.35, bottom=0.35, header=0.15, footer=0.15)
    wb.save(output)
    return output


def _export_audit_operations_basic_xlsx(
    rows: list[dict[str, object]],
    output: Path,
    labels: dict[str, str],
    filters: dict[str, object],
    generated_at: datetime | None,
) -> Path:
    table: list[list[object]] = [
        [labels.get("audit_logs_title", "Operation logs")],
        [labels.get("audit_export_generated_at", "Generated at"), format_datetime_display(generated_at or datetime.now())],
    ]
    if filters.get("data_inicio") or filters.get("start"):
        table.append([labels.get("audit_filter_start", "Start date/time"), filters.get("data_inicio") or filters.get("start")])
    if filters.get("data_fim") or filters.get("end"):
        table.append([labels.get("audit_filter_end", "End date/time"), filters.get("data_fim") or filters.get("end")])
    table.extend([
        [labels.get("audit_export_total", "Total records"), len(rows)],
        [],
        [
            labels.get("audit_col_datetime", "Date/time"),
            labels.get("audit_col_operator", "Operator ID"),
            labels.get("audit_col_tab", "Tab"),
            labels.get("audit_col_action", "Action"),
            labels.get("audit_col_details", "Details"),
        ],
    ])
    for log in format_audit_logs_for_display(rows, labels):
        table.append([
            log.get("data_hora", ""),
            log.get("matricula", ""),
            log.get("aba", ""),
            log.get("acao", ""),
            log.get("detalhes", ""),
        ])

    sheet_rows = []
    for row_number, row in enumerate(table, start=1):
        cells = []
        for column_number, value in enumerate(row, start=1):
            cell_ref = f"{_xlsx_column_name(column_number)}{row_number}"
            text = escape(str(value if value is not None else ""))
            cells.append(f'<c r="{cell_ref}" t="inlineStr"><is><t>{text}</t></is></c>')
        sheet_rows.append(f'<row r="{row_number}">{"".join(cells)}</row>')

    worksheet = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        f'<sheetData>{"".join(sheet_rows)}</sheetData>'
        '</worksheet>'
    )
    with zipfile.ZipFile(output, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr(
            "[Content_Types].xml",
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
            '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
            '<Default Extension="xml" ContentType="application/xml"/>'
            '<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
            '<Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
            '</Types>',
        )
        archive.writestr(
            "_rels/.rels",
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>'
            '</Relationships>',
        )
        archive.writestr(
            "xl/workbook.xml",
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
            'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
            f'<sheets><sheet name="{escape(labels.get("audit_logs_title", "Operation logs")[:31])}" sheetId="1" r:id="rId1"/></sheets>'
            '</workbook>',
        )
        archive.writestr(
            "xl/_rels/workbook.xml.rels",
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>'
            '</Relationships>',
        )
        archive.writestr("xl/worksheets/sheet1.xml", worksheet)
    return output


def _xlsx_column_name(index: int) -> str:
    name = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        name = chr(65 + remainder) + name
    return name


def temperature_status(value: object, low_limit: object, high_limit: object) -> tuple[str, bool]:
    try:
        numeric_value = float(value)
        low = float(low_limit)
        high = float(high_limit)
    except (TypeError, ValueError):
        return "N/A", False
    if low <= numeric_value <= high:
        return "NORMAL / CONFORME", True
    return "ANORMAL / FORA DO PADRAO", False


def temperature_log_filename(identity: ReportIdentity, batch_label: str = "LOTE") -> str:
    pn = clean_identifier(identity.pn)
    sn = "" if is_batch_identifier(identity.pn, batch_label) else identity.sn
    parts = [
        "Log",
        sanitize_path_part(identity.safe_report_number, "-"),
        sanitize_path_part(pn),
        sanitize_path_part(sn),
    ]
    return "_".join(part for part in parts if part) + ".xlsx"


def parts_control_filename(prefix: str = "Controle_Pecas", report_number: object = None, reference_date: datetime | None = None) -> str:
    reference_date = reference_date or datetime.now()
    safe_prefix = sanitize_path_part(prefix) or "Controle_Pecas"
    safe_report = sanitize_path_part(format_report_number(report_number), "-") or "SEM_RELATORIO"
    return f"{safe_prefix}_{safe_report}_{reference_date:%Y-%m-%d}.xlsx"


def parts_control_path(
    root: str | Path,
    report_number: object = None,
    reference_date: datetime | None = None,
    *,
    oven: str | None = None,
    lang: str = "PT",
    prefix: str = "Controle_Pecas",
) -> Path:
    paths = build_output_paths(root, reference_date, oven=oven, lang=lang)
    return paths.entry_control / parts_control_filename(prefix, report_number, reference_date)


def normalize_report_items(items: list[PartItem | dict[str, object] | str]) -> list[PartItem]:
    return [normalize_item(item) for item in items or []]


def report_summary(items: list[PartItem | dict[str, object] | str]) -> dict[str, object]:
    normalized = normalize_report_items(items)
    return {
        "items": normalized,
        "total_quantity": total_quantity(normalized),
        "has_multiple_items": len(normalized) > 1 or any(item.qty > 1 for item in normalized),
    }
